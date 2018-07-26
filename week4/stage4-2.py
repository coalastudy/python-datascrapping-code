
from bs4 import BeautifulSoup
from urllib import request, parse
import openpyxl

# ----------------- 준비 -----------------

dict = {}
keyword = '아시안게임'
period = '7'
# pd: 1일 - 4, 1주 - 1, 1개월 - 2, 6개월 - 6, 1년 - 5, 1시간 - 7, 2시간 - 8, 3시간 - 9, 4시간 - 10, 5시간 - 11, 6시간 - 12

req = request.Request('https://search.naver.com/search.naver?where=news&query=' + parse.quote(keyword) + '&pd=' + period + '&start=1', headers={'User-Agent': 'Mozilla/5.0'})
raw = request.urlopen(req).read()
html = BeautifulSoup(raw, 'html.parser')
total = int(html.select_one('.all_my').text.split('/')[1][:-1].replace(',', '').strip())
page = 1

# ----------------- 수집 -----------------

while (page - 1) * 10 < total:
    req = request.Request('https://search.naver.com/search.naver?where=news&query=' + parse.quote(keyword) + '&pd=' + period + '&start=' + str((page-1) * 10 + 1),
                          headers={'User-Agent': 'Mozilla/5.0'})
    raw = request.urlopen(req).read()
    html = BeautifulSoup(raw, 'html.parser')
    list = html.select('.type01 dl')

    for article in list:
        journal = article.select_one('span._sp_each_source').text
        title = article.select_one('dt').text

        if journal not in dict:
            dict[journal] = [title]
        else:
            if title not in dict[journal]:
                dict[journal].append(title)

    print(page * 10, '/', total)
    page = page + 1

# ----------------- 출력 -----------------

for key in dict.keys():
    print(key, dict[key])

wb = openpyxl.Workbook()
row = 1
ws = wb.active

# ----------------- 저장 -----------------

for journalTitles in dict.items():
    ws.cell(row=row, column=1).value = journalTitles[0]

    for title in journalTitles[1]:
        ws.cell(row=row, column=2).value = title
        row += 1

    row += 1

wb.save(keyword + '.xlsx')