
# import openpyxl

# wb = openpyxl.Workbook()
# wb.save('test.xlsx')


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# sheet['A1'] = 'hello world'
# sheet.cell(row=3, column=3).value = '3, 3'
# sheet.append([1, 2, 3, 4, 5])
#
# wb.save('test2.xlsx')


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet1 = wb['Sheet']
# sheet1.title = '수집 데이터'
# sheet1['A1'] = '첫번째 시트'
#
# sheet2 = wb.create_sheet('정리 결과')
# sheet2['A1'] = '두번째 시트'
#
# sheet1['A2'] = '다시 첫번째 시트'
#
# wb.save('test3.xlsx')


import openpyxl

wb = openpyxl.load_workbook('test2.xlsx')

sheet1 = wb.active

sheet1.title = "이름 변경"
sheet1.append(range(10))

wb.save('test2.xlsx')

