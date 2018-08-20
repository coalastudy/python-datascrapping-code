import requests
from bs4 import BeautifulSoup
import openpyxl

journal_articles = {}

keyword = input('검색어를 입력해주세요 : ')

for i in range(3):
    raw = requests.get('https://search.naver.com/search.naver?&where=news&query=' + keyword + '&start=' + str(i * 10 + 1), headers={'User-Agent': 'Mozilla/5.0'}).text
    html = BeautifulSoup(raw, 'html.parser')

    list = html.select('.type01 > li')

    for article in list:
        journal = article.select_one('span._sp_each_source').text
        title = article.select_one('a._sp_each_title').text

        if journal in journal_articles.keys():
            journal_articles[journal].append(title)
        else:
            journal_articles[journal] = [title]

for item in journal_articles.items():
    print(item)

wb = openpyxl.Workbook()
row = 1
ws = wb.active

for journal_titles in journal_articles.items():
    ws.cell(row=row, column=1).value = journal_titles[0]

    for title in journal_titles[1]:
        ws.cell(row=row, column=2).value = title
        row += 1

    row += 1

wb.save(keyword + '.xlsx')
